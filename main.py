import os
import re
import sys
import time

import graphviz
import xlrd
import xlsxwriter
import openpyxl
from networkx.drawing.nx_agraph import read_dot
from unidiff import PatchSet
from openpyxl.utils import get_column_letter

from memory_leak_analyzer import MemoryLeakAnalyzer
from static_analyze_util import get_nodes_to_file, is_node_in_file
from use_after_free_analyzer import UseAfterFreeAnalyzer
from diff_directories import generate_diff_by_path


def get_patch_info(file_path):
    with open(file_path, 'r') as patch_file:
        patch_content = patch_file.read()
    patch_info = PatchSet(patch_content)
    return patch_info


# 判断是否是注释行
def is_commnet(line):
    return re.match(r'^\s*(//)|(^\s*/\*)', line) is not None


# 获取patchinfo中增加的line和减少的line
def get_added_removed_line(patch_info):
    added_line_info = {}
    removed_line_info = {}
    for patched_file in patch_info:
        # 删除或者新增文件场景，行为比较敏感，足够引起审核者的注意，因此这里不进行分析
        if patched_file.is_removed_file or patched_file.is_added_file:
            continue
        else:
            added_line = []
            removed_line = []
            source_file = patched_file.source_file
            target_file = patched_file.target_file

            for hunk in patched_file:
                target_lines = list(hunk.target_lines())
                source_lines = list(hunk.source_lines())
                for line in target_lines:
                    if (not is_commnet(line.value)) and line.is_added:
                        added_line.append(line)
                        # print(f"added line is {line.value.strip()}")
                for line in source_lines:
                    if (not is_commnet(line.value)) and line.is_removed:
                        removed_line.append(line)
                        # print(f"removed line is {line.value.strip()}")
            if bool(added_line):
                added_line_info[target_file] = added_line
            if bool(removed_line):
                removed_line_info[source_file] = removed_line
    return added_line_info, removed_line_info


# 通过joern生成待测源码的cpg
def generate_cpg(code_path):
    if os.path.exists(code_path + '/a/outA'):
        os.system(f"sudo rm -rf {code_path}/a/outA")
    if os.path.exists(code_path + '/b/outB'):
        os.system(f"sudo rm -rf {code_path}/b/outB")

    os.system(f"cd /home/tbx/bin/joern/joern-cli/; sudo ./joern-parse {code_path}/a;"
              f" sudo ./joern-export --repr=all --format=dot --out {code_path}/a/outA")
    os.system(f"cd /home/tbx/bin/joern/joern-cli/; sudo ./joern-parse {code_path}/b;"
              f" sudo ./joern-export --repr=all --format=dot --out {code_path}/b/outB")


def update_result(case_path: str, vul_type: str):
    results_file = 'results/results.xlsx'
    # 如果case_path不存在则新建一行，并先都填入0
    if os.path.exists(results_file):
        # 根据漏洞类型，在对应列填入1
        workbook = openpyxl.load_workbook(results_file)
        worksheet = workbook.active

        is_duplicate = False
        duplicate_row = None

        # 从第三行开始读取，
        last_row = 2
        for row_index in range(3, worksheet.max_row + 1):
            cell_value = worksheet.cell(row=row_index, column=1).value
            if cell_value is None or cell_value == "":
                break
            # 如果测试用例已经存在，则直接更新，否则新增一行
            if str(cell_value).strip() == case_path.strip():
                is_duplicate = True
                duplicate_row = row_index
                break

            last_row = row_index

        # 如果没有记录，则新增一行
        if not is_duplicate:
            next_row = last_row + 1
            worksheet.cell(row=next_row, column=1).value = case_path

            for col_index in range(2, 6):
                worksheet.cell(row=next_row, column=col_index).value = 0
            if vul_type == 'UAF':
                worksheet.cell(row=next_row, column=2).value = 1
            if vul_type == 'MemoryLeak':
                worksheet.cell(row=next_row, column=4).value = 1
            workbook.save(results_file)
        else:
            if vul_type == 'UAF':
                worksheet.cell(row=duplicate_row, column=2).value = 1
            if vul_type == 'MemoryLeak':
                worksheet.cell(row=duplicate_row, column=4).value = 1
            workbook.save(results_file)


def analyze_graph(lines, dot_file_path, file_type, file_name, case_path):
    risk_set = set()
    if len(lines) < 0:
        return
    graph = read_dot(dot_file_path)
    # print(f"Test dot_file_path is {dot_file_path}")
    # 创建缺陷解析器实例
    memory_leak_analyzer = MemoryLeakAnalyzer(graph)
    # strict：严格模式，只返回触发UAF的风险，非严格模式则会一并统计空指针/未成熟漏洞等风险
    use_after_free_analyzer = UseAfterFreeAnalyzer(graph, False)

    # 获取node和filename之间的映射关系
    nodes_to_file = get_nodes_to_file(graph)
    # 为了方便后续查询，创建一个line_num和nodes的映射，后续查询可以通过matching_nodes = nodes_line_num.get(target_line, [])
    nodes_line_num = {}
    for node, data in graph.nodes(data=True):
        # print(is_node_in_file(nodes_to_file, node, file_name))
        if 'LINE_NUMBER' in data and is_node_in_file(nodes_to_file, node, file_name):
            line_num = int(data['LINE_NUMBER'])
            if line_num not in nodes_line_num:  # and filename == node file
                nodes_line_num[line_num] = []
            nodes_line_num[line_num].append((node, data))
    # 针对修改逐行分析
    # print(f"Test nodes line num size is {len(nodes_line_num)}")
    for line in lines:
        # 找到图中对应的节点集合，matching_nodes
        matching_nodes = list()
        if file_type == "source":
            # print(f"Test source: Line value is {line.value.strip()}, line num is {line.source_line_no}")
            matching_nodes = nodes_line_num.get(line.source_line_no, [])
        elif file_type == "target":
            # print(f"Test target: Line value is {line.value.strip()}, line num is {line.target_line_no}")
            matching_nodes = nodes_line_num.get(line.target_line_no, [])

        if not matching_nodes:
            print(f"There is no matching nodes in CPG, line value is '{line.value.strip()}'")
        else:
            # print(f"Test: the node contained line num {line.target_line_no} are {matching_nodes}")
            # 可能存在同时删除 内存分配和内存释放 的情况，我们并不做过滤，这种简单的判断留给committer
            # 1.检测内存泄漏风险
            ml_start_time = time.time()
            risk, res_str = memory_leak_analyzer.analyze_potential_leak(matching_nodes, line, file_type)
            ml_end_time = time.time()
            print(f'Memory leak analysis for a single line takes time: {(ml_end_time - ml_start_time) * 1000:.3f} ms')
            # print(f"Test res_str of memory leak analyze is {res_str}")
            if risk:
                risk_set.add(res_str)
                update_result(case_path, 'MemoryLeak')
            # 2.检测UAF风险
            uaf_start_time = time.time()
            risk, res_str = use_after_free_analyzer.analyze_potential_uaf(matching_nodes, line, file_type)
            uaf_end_time = time.time()
            print(f'UAF analysis for a single line takes time: {(uaf_end_time - uaf_start_time) * 1000:.3f} ms')
            # print(f"Test res_str of use after free analyze is {res_str}")
            if risk:
                risk_set.add(res_str)
                update_result(case_path, 'UAF')
    return risk_set


# 根据dot文件生成图片，测试使用
def dot_to_image(file_path, out_put_path, format='png'):
    with open(file_path, 'r') as file:
        dot_graph = file.read()
    graph = graphviz.Source(dot_graph)

    graph.render(out_put_path, format=format, cleanup=True)
    print(f"Image saved as {out_put_path}.{format}")


# 递归删除某个名称的目录
def delete_folders(root_path, folder_names):
    if not os.path.exists(root_path) or not os.path.isdir(root_path):
        print(f"Error: Path '{root_path}' does not exist or is not a directory.")
        return 0

    count = 0
    # 获取当前层级的所有子目录
    items = os.listdir(root_path)
    dirs_to_process = []

    # 先检查当前层级是否有要删除的文件夹
    for item in items:
        item_path = os.path.join(root_path, item)

        if os.path.isdir(item_path):
            if item in folder_names:
                print(f"Remove directory: {item_path}")
                try:
                    # shutil.rmtree(item_path)
                    os.system(f"sudo rm -rf {item_path}")
                    count += 1
                except Exception as e:
                    print(f"Remove directory '{item_path}' error: {e}.")
            else:
                # 将不需要删除的子目录添加到待处理列表
                dirs_to_process.append(item_path)

    # 递归处理剩余子目录
    for dir_path in dirs_to_process:
        count += delete_folders(dir_path, folder_names)

    return count


def data_preprocess(case_path: str, mk_new_diff: bool, mk_new_cpg: bool) -> str:
    # data_path 是test_case的路径
    patch_path = generate_diff_by_path(mk_new_diff, case_path)

    if patch_path != '' and mk_new_cpg:
        start_time = time.time()
        generate_cpg(case_path)
        end_time = time.time()
        print(f'Time to create CPG: {(end_time - start_time) * 1000:.3f} ms')

    return patch_path


def get_subdirectories(path):
    if not os.path.exists(path):
        print(f'Path {path} does not exist.')
        return set()

    if not os.path.isdir(path):
        print(f'Path {path} is not directory.')
        return set()

    subdirectories = set()
    for item in os.listdir(path):
        item_path = os.path.join(path, item)
        if os.path.isdir(item_path):
            subdirectories.add(item)

    return subdirectories


def execute_by_path(test_case_path: str, is_root_path: bool, mk_new_diff: bool, mk_new_cpg: bool):
    if not is_root_path:
        patch_path = data_preprocess(test_case_path, mk_new_diff, mk_new_cpg)
        if patch_path == '':
            print(f"Data preprocess fail, can not generate new patch...")
            sys.exit(0)

        patch_info = get_patch_info(patch_path)
        source_cpg_path = test_case_path + "/a/outA/export.dot"
        target_cpg_path = test_case_path + "/b/outB/export.dot"

        # 查看图中路径，注意非常耗时
        # dot_to_image(target_cpg_path, code_path+'/b/cpg_img', 'png')
        added_line_info, removed_line_info = get_added_removed_line(patch_info)

        # 如果是added_line，则对target文件（b文件），如果是removed_line则对source文件进行分析
        if bool(added_line_info):
            # 针对增加场景，分析提交后的版本
            start_time = time.time()
            for file_name, added_lines in added_line_info.items():
                risk_set = analyze_graph(added_lines, target_cpg_path, "target", file_name, test_case_path)
                if risk_set:
                    print("The potential risks by the added lines in this commit are: ")
                    for risk in risk_set:
                        print(f"#### {risk} ####")
                else:
                    print("#### Adding in this commit is safe ####")
            end_time = time.time()
            print(f'Patch defect analysis takes time: {(end_time - start_time) * 1000:.3f} ms')
        if bool(removed_line_info):
            # 针对删除行场景，我们分析提交前的版本
            start_time = time.time()
            for file_name, removed_lines in removed_line_info.items():
                risk_set = analyze_graph(removed_lines, source_cpg_path, "source", file_name, test_case_path)
                # 后面可以考虑生成文本文件，然后输出到对应patch的目录下
                if risk_set:
                    print("The potential risks by the removed lines in this commit are: ")
                    for risk in risk_set:
                        print(f"#### {risk} ####")
                else:
                    print("#### Removing in this commit is safe ####")
            end_time = time.time()
            print(f'Patch defect analysis takes time: {(end_time - start_time) * 1000:.3f} ms')
    else:
        # 注意！根目录批量实验只针对UAF/MemoryLeak两个目录，因为目录遍历规则是固定的
        if not test_case_path.endswith('UAF') and not test_case_path.endswith('MemoryLeak'):
            print(f'Error path! batch testing is only supported for the UAF and MemoryLeak directories.')
        else:
            bad_path = test_case_path + "/bad"
            good_path = test_case_path + "/good"

            bad_path_add = bad_path + "/Addition"
            bad_path_remove = bad_path + "/Removement"

            good_path_add = good_path + "/Addition"
            good_path_remove = good_path + "/Removement"

            cases_in_bad_path_add = get_subdirectories(bad_path_add)
            cases_in_bad_path_remove = get_subdirectories(bad_path_remove)

            cases_in_good_path_add = get_subdirectories(good_path_add)
            cases_in_good_path_remove = get_subdirectories(good_path_remove)

            print(f'Cases in {bad_path_add} are:')
            for tmp_dire in sorted(cases_in_bad_path_add):
                case_dire = bad_path_add + '/' + tmp_dire
                print(f"- {case_dire}")
                # 遍历子目录，然后调用execute_by_path(directory, False, mk_new_diff, mk_new_cpg)
                execute_by_path(case_dire, False, mk_new_diff, mk_new_cpg)

            print(f'Cases in {bad_path_remove} are:')
            for tmp_dire in sorted(cases_in_bad_path_remove):
                case_dire = bad_path_remove + '/' + tmp_dire
                print(f"- {case_dire}")
                # 遍历子目录，然后调用execute_by_path(directory, False, mk_new_diff, mk_new_cpg)
                execute_by_path(case_dire, False, mk_new_diff, mk_new_cpg)

            print(f'Cases in {good_path_add} are:')
            for tmp_dire in sorted(cases_in_good_path_add):
                case_dire = good_path_add + '/' + tmp_dire
                print(f"- {case_dire}")
                # 遍历子目录，然后调用execute_by_path(directory, False, mk_new_diff, mk_new_cpg)
                execute_by_path(case_dire, False, mk_new_diff, mk_new_cpg)

            print(f'Cases in {good_path_remove} are:')
            for tmp_dire in sorted(cases_in_good_path_remove):
                case_dire = good_path_remove + '/' + tmp_dire
                print(f"- {case_dire}")
                # 遍历子目录，然后调用execute_by_path(directory, False, mk_new_diff, mk_new_cpg)
                execute_by_path(case_dire, False, mk_new_diff, mk_new_cpg)


def clear_intermediate_files(path: str):
    delete_folders(path, 'outA')
    delete_folders(path, 'outB')
    delete_folders(path, 'patch-info')


def init_result_file():
    re_file_path = 'results/results.xlsx'
    # 如果文件不存在则复制一个文件
    if os.path.exists(re_file_path):
        print(f'The file {re_file_path} already exists and does not need to be created.')
        return

    try:
        os.system('cp results/results_original.xlsx results/results.xlsx')
    except Exception as e:
        print(f'Error: {str(e)}')


if __name__ == '__main__':
    # 初始化结果文件
    init_result_file()
    # 对测试用例的数据进行预处理，输入为测试用例目录，创建diff文件，生成patch文件和cpg文件
    test_case_path = "/home/tbx/workspace/DataSet-2022-08-11-juliet/UAF"
    # test_case_path = "/home/tbx/workspace/DataSet-2022-08-11-juliet/UAF/bad/Removement/testcase_17"
    # test_case_path = "/home/tbx/workspace/DataSet-2022-08-11-juliet/Hypocrite-Commit/case_3"
    # test_case_path = "/home/tbx/workspace/DataSet-2022-08-11-juliet/CVE-2019012819"

    # 针对UAF和MemoryLeak下的批量实验，is_root_path为True；单个case实验is_root_path为False
    execute_by_path(test_case_path, True, True, True)

    # 按需删除产生的outA,outB和patch-info
    # clear_intermediate_files(test_case_path)
